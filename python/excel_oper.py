'''
Author: 王政乔 me@zhengqiao.wang
Date: 2022-09-30 19:29:23
LastEditors: 王政乔 me@zhengqiao.wang
LastEditTime: 2022-09-30 19:49:53
FilePath: /codesnippet/python/excel_oper.py
Description: Excel操作，可以用于简单的Excel读取，方便广大办公人员实现自动化办公
Website: https://www.zhengqiao.wang
'''

import openpyxl


class ExcelOper:
    """Excel操作
    可以解析Microsoft Excel 2003和2007的数据，并读取其中的指定列
    """

    def __init__(self, file_path, sheet_idx) -> None:
        self.__file_path = file_path
        self.__col_idx = {}
        self.__start_idx = 1
        self.__end_idx = 1
        self.__sheet_idx = sheet_idx

    def setStartLine(self, idx):
        """设置读取起始行

        Args:
            idx (int): 起始行，直接对应Excel表中行号，如果是负数则取该表倒数第n行

        Returns:
            ExcelOper: 返回对象自身
        """
        self.__start_idx = idx
        return self

    def setEndLine(self, idx):
        """设置读取终止行

        Args:
            idx (int): 起始行，直接对应Excel表中行号，如果是负数则取该表倒数第n行

        Returns:
            ExcelOper: 返回对象自身
        """
        self.__end_idx = idx
        return self

    def registColGet(self, col_name, col_idx):
        """注册要获取的列

        Args:
            col_name (str): 列名，唯一，返回的数据中以该列名为key
            col_idx (int): 列序号，A对应1，B对应2....以此类推

        Returns:
            ExcelOper: 返回对象自身
        """
        self.__col_idx[col_name] = col_idx
        return self

    def getDatas(self):
        """读取数据

        Returns:
            generator: 以列名为key，以值为value的dict的list

        Yields:
            dict: 以列名为key，以值为value的dict
        """
        self.__book = openpyxl.load_workbook(self.__file_path)
        if self.__book == None:
            return None
        self.__sheet = self.__book.worksheets[self.__sheet_idx]
        if self.__sheet == None:
            return None
        self.__max_row = self.__sheet.max_row
        if self.__start_idx < 0:
            self.__start_idx = self.__max_row + self.__start_idx + 1
        if self.__end_idx < 0:
            self.__end_idx = self.__max_row + self.__end_idx + 1
        for i in range(self.__start_idx, self.__end_idx + 1):
            tmp_dict = {}
            for col_name, col_idx in self.__col_idx.items():
                tmp_dict[col_name] = self.__sheet.cell(
                    row=i, column=col_idx).value
            yield tmp_dict

    @staticmethod
    def translateToXLSX(file_name):
        """将xls转换为xlsx。将会在本地新建一个.xlsx文件。

        Args:
            file_name (str): xls文件路径

        Raises:
            RuntimeError: 无法转换

        Returns:
            str: 转换后的xlsx文件地址
        """
        if file_name.split(".")[-1] == "xlsx":
            return file_name
        elif file_name.split(".")[-1] == "xls":
            # excel 2003
            file_name = file_name.replace("\\\\", "\\")
            new_file_name = file_name+"x"
            print(file_name, "->", new_file_name)
            import win32com.client as win32
            excel = win32.gencache.EnsureDispatch('Excel.Application')
            wb = excel.Workbooks.Open(file_name)
            wb.SaveAs(new_file_name, FileFormat=51)
            wb.Close()
            excel.Application.Quit()
            return new_file_name
        else:
            raise RuntimeError("无法识别的excel文件后缀，请确认文件后缀是xls或xlsx")
